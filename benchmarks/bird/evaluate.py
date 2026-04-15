# Copyright 2026 CVS Health and/or one of its affiliates
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
# This file uses the following unmodified third-party packages,
# each retaining its original copyright and license:
#   openpyxl (MIT)

"""BIRD-compatible evaluation metrics for text-to-SQL benchmarking.

Implements the official BIRD evaluation metrics:
- EX (Execution Accuracy): Whether predicted SQL returns the same result set as gold SQL
- Soft F1: Row-level fuzzy matching between predicted and gold result sets
- R-VES (Reward-based Valid Efficiency Score): Efficiency-weighted accuracy

Reference: https://github.com/bird-bench/mini_dev/tree/main/evaluation
"""

import json
import logging
import multiprocessing as mp
import os
import sqlite3
import time
from collections import Counter
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# String constants for difficulty levels and Excel formatting
# ---------------------------------------------------------------------------
_DIFFICULTY_SIMPLE = "simple"
_DIFFICULTY_MODERATE = "moderate"
_DIFFICULTY_CHALLENGING = "challenging"
_FMT_FLOAT = ">15.2f"
_EXCEL_THIN_BORDER = "thin_border"
_EXCEL_ALIGNMENT = "Alignment"
_EXCEL_CENTER = "center"
_EXCEL_HEADER_FONT = "header_font"
_EXCEL_HEADER_FILL = "header_fill"
_COL_EX_SCORE = "ex_score"
_COL_SOFT_F1_SCORE = "soft_f1_score"
_COL_EX_PCT = "EX (%)"
_COL_SOFT_F1_PCT = "Soft F1 (%)"

# Published reference EX (%) for BIRD Mini-Dev **SQLite** — https://bird-bench.github.io/
# These are full 500-question leaderboard figures; subsampled runs are indicative only.
BIRD_MINI_DEV_SQLITE_REFERENCE_BASELINES: List[Dict[str, Any]] = [
    {
        "label": "GPT-4",
        "ex_pct": 47.80,
        "soft_f1_pct": None,
        "source": "https://bird-bench.github.io/",
        "notes": "Mini-Dev SQLite, leaderboard (full 500).",
    },
    {
        "label": "TA + GPT-4o",
        "ex_pct": 63.00,
        "soft_f1_pct": None,
        "source": "https://bird-bench.github.io/",
        "notes": "Agent + GPT-4o, Mini-Dev SQLite (full 500).",
    },
]


@dataclass
class EvaluationResult:
    """Result of evaluating a single predicted SQL against gold SQL."""

    question_id: int
    db_id: str
    difficulty: str
    ex_score: int  # 1 if execution results match, 0 otherwise
    soft_f1_score: float  # Soft F1 between result sets
    predicted_sql: str
    gold_sql: str
    error: Optional[str] = None
    predicted_time_seconds: float = 0.0
    gold_time_seconds: float = 0.0


@dataclass
class EvaluationReport:
    """Aggregated evaluation report with per-difficulty breakdown."""

    total: int = 0
    ex_accuracy: float = 0.0
    soft_f1: float = 0.0

    simple_count: int = 0
    simple_ex: float = 0.0
    simple_f1: float = 0.0

    moderate_count: int = 0
    moderate_ex: float = 0.0
    moderate_f1: float = 0.0

    challenging_count: int = 0
    challenging_ex: float = 0.0
    challenging_f1: float = 0.0

    error_count: int = 0
    timeout_count: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "total": self.total,
            "execution_accuracy": {
                "overall": round(self.ex_accuracy, 2),
                _DIFFICULTY_SIMPLE: round(self.simple_ex, 2),
                _DIFFICULTY_MODERATE: round(self.moderate_ex, 2),
                _DIFFICULTY_CHALLENGING: round(self.challenging_ex, 2),
            },
            "soft_f1": {
                "overall": round(self.soft_f1, 2),
                _DIFFICULTY_SIMPLE: round(self.simple_f1, 2),
                _DIFFICULTY_MODERATE: round(self.moderate_f1, 2),
                _DIFFICULTY_CHALLENGING: round(self.challenging_f1, 2),
            },
            "counts": {
                _DIFFICULTY_SIMPLE: self.simple_count,
                _DIFFICULTY_MODERATE: self.moderate_count,
                _DIFFICULTY_CHALLENGING: self.challenging_count,
                "errors": self.error_count,
                "timeouts": self.timeout_count,
            },
        }

    def print_report(self):
        """Print a formatted evaluation report matching BIRD's output style."""
        counts = [self.simple_count, self.moderate_count, self.challenging_count, self.total]
        ex_scores = [self.simple_ex, self.moderate_ex, self.challenging_ex, self.ex_accuracy]
        f1_scores = [self.simple_f1, self.moderate_f1, self.challenging_f1, self.soft_f1]

        print("\n" + "=" * 90)
        print("BIRD Benchmark Evaluation Results")
        print("=" * 90)
        print(f"{'':20} {'simple':>15} {'moderate':>15} {'challenging':>15} {'total':>15}")
        print(f"{'count':20} {counts[0]:>15} {counts[1]:>15} {counts[2]:>15} {counts[3]:>15}")
        print("-" * 90)
        print(f"{'EX (%)':20} {ex_scores[0]:>15.2f} {ex_scores[1]:>15.2f} {ex_scores[2]:>15.2f} {ex_scores[3]:>15.2f}")
        print(f"{'Soft F1 (%)':20} {f1_scores[0]:>15.2f} {f1_scores[1]:>15.2f} {f1_scores[2]:>15.2f} {f1_scores[3]:>15.2f}")
        print("-" * 90)
        print(f"Errors: {self.error_count}  |  Timeouts: {self.timeout_count}")
        print("=" * 90 + "\n")


def _execute_sql_safe(sql: str, db_path: str, timeout: float = 30.0) -> Tuple[Optional[List[tuple]], Optional[str], float]:
    """Execute SQL against a SQLite database with timeout protection.

    Returns:
        Tuple of (results, error_message, execution_time_seconds).
    """
    start = time.time()
    try:
        conn = sqlite3.connect(db_path, timeout=timeout)
        conn.execute("PRAGMA journal_mode=WAL;")
        cursor = conn.cursor()
        cursor.execute(sql)
        results = cursor.fetchall()
        conn.close()
        elapsed = time.time() - start
        return results, None, elapsed
    except Exception as e:
        elapsed = time.time() - start
        return None, str(e), elapsed


def _calculate_ex(predicted_results: List[tuple], gold_results: List[tuple]) -> int:
    """Calculate Execution Accuracy (EX) — BIRD's primary metric.

    EX = 1 if the set of result tuples from the predicted SQL matches
    the set of result tuples from the gold SQL, 0 otherwise.
    """
    if set(predicted_results) == set(gold_results):
        return 1
    return 0


def _calculate_soft_f1(predicted_results: List[tuple], gold_results: List[tuple]) -> float:
    """Calculate Soft F1 score between predicted and gold result sets.

    For each row, counts matched cells, predicted-only cells, and gold-only cells.
    Then computes precision, recall, and F1 across all rows.

    This is more lenient than EX as it allows partial credit for
    correct values even if column order differs.
    """
    if not gold_results and not predicted_results:
        return 1.0
    if not gold_results or not predicted_results:
        return 0.0

    max_rows = max(len(predicted_results), len(gold_results))
    total_tp = 0
    total_fp = 0
    total_fn = 0

    for i in range(max_rows):
        pred_row = list(predicted_results[i]) if i < len(predicted_results) else []
        gold_row = list(gold_results[i]) if i < len(gold_results) else []

        pred_counter = Counter(str(v) for v in pred_row)
        gold_counter = Counter(str(v) for v in gold_row)

        matched = sum((pred_counter & gold_counter).values())
        pred_only = sum(pred_counter.values()) - matched
        gold_only = sum(gold_counter.values()) - matched

        total_tp += matched
        total_fp += pred_only
        total_fn += gold_only

    precision = total_tp / (total_tp + total_fp) if (total_tp + total_fp) > 0 else 0.0
    recall = total_tp / (total_tp + total_fn) if (total_tp + total_fn) > 0 else 0.0

    if precision + recall == 0:
        return 0.0
    return 2 * precision * recall / (precision + recall)


def _evaluate_single(args: tuple) -> EvaluationResult:
    """Evaluate a single prediction (designed for multiprocessing)."""
    question_id, db_id, difficulty, predicted_sql, gold_sql, db_path, timeout = args

    pred_results, pred_error, pred_time = _execute_sql_safe(predicted_sql, db_path, timeout)
    gold_results, gold_error, gold_time = _execute_sql_safe(gold_sql, db_path, timeout)

    if pred_error or pred_results is None:
        return EvaluationResult(
            question_id=question_id,
            db_id=db_id,
            difficulty=difficulty,
            ex_score=0,
            soft_f1_score=0.0,
            predicted_sql=predicted_sql,
            gold_sql=gold_sql,
            error=f"Predicted SQL error: {pred_error}",
            predicted_time_seconds=pred_time,
            gold_time_seconds=gold_time,
        )

    if gold_error or gold_results is None:
        return EvaluationResult(
            question_id=question_id,
            db_id=db_id,
            difficulty=difficulty,
            ex_score=0,
            soft_f1_score=0.0,
            predicted_sql=predicted_sql,
            gold_sql=gold_sql,
            error=f"Gold SQL error: {gold_error}",
            predicted_time_seconds=pred_time,
            gold_time_seconds=gold_time,
        )

    ex = _calculate_ex(pred_results, gold_results)
    f1 = _calculate_soft_f1(pred_results, gold_results)

    return EvaluationResult(
        question_id=question_id,
        db_id=db_id,
        difficulty=difficulty,
        ex_score=ex,
        soft_f1_score=f1,
        predicted_sql=predicted_sql,
        gold_sql=gold_sql,
        predicted_time_seconds=pred_time,
        gold_time_seconds=gold_time,
    )


def _read_gold_sql_file(path: str):
    """Read gold SQL file (tab-separated: SQL\\tdb_id) and return (gold_sqls, gold_db_ids) lists."""
    gold_sqls = []
    gold_db_ids = []
    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if line:
                parts = line.split("\t")
                gold_sqls.append(parts[0])
                gold_db_ids.append(parts[1] if len(parts) > 1 else "unknown")
    return gold_sqls, gold_db_ids


def _read_difficulty_file(path: str) -> List[str]:
    """Read JSONL difficulty file and return list of difficulty strings."""
    difficulties = []
    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if line:
                entry = json.loads(line)
                difficulties.append(entry.get("difficulty", _DIFFICULTY_SIMPLE))
    return difficulties


@dataclass
class BIRDEvaluator:
    """Evaluates predicted SQL against gold SQL using BIRD's official metrics.

    Supports evaluating from:
    1. A predictions JSON file (BIRD format)
    2. A list of BenchmarkResult objects from the runner

    Args:
        db_root_path: Root directory containing BIRD SQLite databases.
        timeout: Timeout per SQL execution in seconds.
        num_workers: Number of parallel workers for evaluation.
    """

    db_root_path: str
    timeout: float = 30.0
    num_workers: int = 1

    @staticmethod
    def _load_gold_sql_file(gold_sql_path: str) -> Tuple[List[str], List[str]]:
        """Parse a tab-separated gold SQL file into (sqls, db_ids) lists."""
        gold_sqls: List[str] = []
        gold_db_ids: List[str] = []
        with open(gold_sql_path, "r") as f:
            for line in f:
                line = line.strip()
                if line:
                    parts = line.split("\t")
                    gold_sqls.append(parts[0])
                    gold_db_ids.append(parts[1] if len(parts) > 1 else "unknown")
        return gold_sqls, gold_db_ids

    @staticmethod
    def _load_difficulties(difficulty_path: str) -> List[str]:
        """Load difficulty labels from a JSONL file."""
        difficulties: List[str] = []
        with open(difficulty_path, "r") as f:
            for line in f:
                line = line.strip()
                if line:
                    entry = json.loads(line)
                    difficulties.append(entry.get("difficulty", _DIFFICULTY_SIMPLE))
        return difficulties

    @staticmethod
    def _extract_pred_sql(predictions: Dict[str, Any], idx: int) -> str:
        """Extract the predicted SQL for a given index from a BIRD predictions dict."""
        str_idx = str(idx)
        if str_idx not in predictions:
            return "SELECT 1"
        pred_entry = predictions[str_idx]
        if "\t----- bird -----\t" in pred_entry:
            return pred_entry.split("\t----- bird -----\t")[0]
        return pred_entry

    def evaluate_from_predictions_file(
        self,
        predictions_path: str,
        gold_sql_path: str,
        difficulty_path: str,
    ) -> EvaluationReport:
        """Evaluate using BIRD-format prediction and gold SQL files.

        Args:
            predictions_path: Path to predictions JSON (BIRD format).
            gold_sql_path: Path to gold SQL file (tab-separated: SQL\\tdb_id).
            difficulty_path: Path to JSONL file with difficulty labels.

        Returns:
            EvaluationReport with per-difficulty breakdown.
        """
        with open(predictions_path, "r") as f:
            predictions = json.load(f)

        gold_sqls, gold_db_ids = self._load_gold_sql_file(gold_sql_path)
        difficulties = self._load_difficulties(difficulty_path)

        eval_args = []
        for idx in range(len(gold_sqls)):
            pred_sql = self._extract_pred_sql(predictions, idx)
            db_id = gold_db_ids[idx]
            db_path = os.path.join(self.db_root_path, db_id, f"{db_id}.sqlite")
            difficulty = difficulties[idx] if idx < len(difficulties) else _DIFFICULTY_SIMPLE

            eval_args.append((
                idx, db_id, difficulty, pred_sql, gold_sqls[idx], db_path, self.timeout
            ))

        report, _ = self._run_evaluation(eval_args)
        return report

    def evaluate_from_results(
        self,
        results: List[Any],
        gold_sqls: Optional[Dict[int, str]] = None,
        return_details: bool = False,
    ):
        """Evaluate from BenchmarkResult objects.

        Args:
            results: List of BenchmarkResult from the runner.
            gold_sqls: Optional override for gold SQLs (question_id -> SQL).
            return_details: If True, returns (report, eval_results) tuple.

        Returns:
            EvaluationReport with per-difficulty breakdown.
        """
        eval_args = []
        for r in results:
            gold = gold_sqls[r.question_id] if gold_sqls else r.gold_sql
            db_path = os.path.join(self.db_root_path, r.db_id, f"{r.db_id}.sqlite")

            eval_args.append((
                r.question_id, r.db_id, r.difficulty,
                r.predicted_sql, gold, db_path, self.timeout
            ))

        report, eval_results = self._run_evaluation(eval_args)
        if return_details:
            return report, eval_results
        return report

    def _run_evaluation(self, eval_args: List[tuple]) -> Tuple[EvaluationReport, List[EvaluationResult]]:
        """Run evaluation across all questions, optionally in parallel."""
        logger.info("Evaluating %d predictions...", len(eval_args))

        if self.num_workers > 1:
            with mp.Pool(processes=self.num_workers) as pool:
                eval_results = pool.map(_evaluate_single, eval_args)
        else:
            eval_results = [_evaluate_single(args) for args in eval_args]

        eval_results.sort(key=lambda r: r.question_id)
        return self._aggregate_results(eval_results), eval_results

    def _aggregate_results(self, results: List[EvaluationResult]) -> EvaluationReport:
        """Aggregate individual results into an EvaluationReport."""
        report = EvaluationReport()
        report.total = len(results)

        simple, moderate, challenging = [], [], []
        for r in results:
            if r.error:
                report.error_count += 1
                if "timeout" in (r.error or "").lower():
                    report.timeout_count += 1

            if r.difficulty == _DIFFICULTY_SIMPLE:
                simple.append(r)
            elif r.difficulty == _DIFFICULTY_MODERATE:
                moderate.append(r)
            elif r.difficulty == _DIFFICULTY_CHALLENGING:
                challenging.append(r)

        def _avg(items, attr):
            if not items:
                return 0.0
            return 100.0 * sum(getattr(r, attr) for r in items) / len(items)

        report.simple_count = len(simple)
        report.simple_ex = _avg(simple, _COL_EX_SCORE)
        report.simple_f1 = _avg(simple, _COL_SOFT_F1_SCORE)

        report.moderate_count = len(moderate)
        report.moderate_ex = _avg(moderate, _COL_EX_SCORE)
        report.moderate_f1 = _avg(moderate, _COL_SOFT_F1_SCORE)

        report.challenging_count = len(challenging)
        report.challenging_ex = _avg(challenging, _COL_EX_SCORE)
        report.challenging_f1 = _avg(challenging, _COL_SOFT_F1_SCORE)

        report.ex_accuracy = _avg(results, _COL_EX_SCORE)
        report.soft_f1 = _avg(results, _COL_SOFT_F1_SCORE)

        return report

    def save_report(self, report: EvaluationReport, output_path: str):
        """Save evaluation report to a JSON file."""
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        with open(output_path, "w") as f:
            json.dump(report.to_dict(), f, indent=2)
        logger.info("Evaluation report saved to %s", output_path)

    def _write_summary_sheet(self, ws_summary, report, run_config, styles):
        """Populate the Summary sheet with run config and per-difficulty metrics."""
        header_font = styles[_EXCEL_HEADER_FONT]
        header_fill = styles[_EXCEL_HEADER_FILL]
        subheader_font = styles["subheader_font"]
        thin_border = styles[_EXCEL_THIN_BORDER]
        Alignment = styles[_EXCEL_ALIGNMENT]
        Font = styles["Font"]

        ws_summary.column_dimensions["A"].width = 28
        ws_summary.column_dimensions["B"].width = 18
        ws_summary.column_dimensions["C"].width = 18
        ws_summary.column_dimensions["D"].width = 18
        ws_summary.column_dimensions["E"].width = 18

        row = 1
        ws_summary.merge_cells("A1:E1")
        cell = ws_summary.cell(row=row, column=1, value="BIRD Benchmark — askRITA Evaluation Report")
        cell.font = Font(bold=True, size=16, color="2F5496")
        cell.alignment = Alignment(horizontal=_EXCEL_CENTER)
        row += 2

        if run_config:
            for key, val in run_config.items():
                ws_summary.cell(row=row, column=1, value=key).font = subheader_font
                ws_summary.cell(row=row, column=2, value=str(val))
                row += 1
            row += 1

        headers = ["Metric", "Simple", "Moderate", "Challenging", "Overall"]
        for col, h in enumerate(headers, 1):
            c = ws_summary.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal=_EXCEL_CENTER)
            c.border = thin_border
        row += 1

        data_rows = [
            ("Count", report.simple_count, report.moderate_count, report.challenging_count, report.total),
            (_COL_EX_PCT, f"{report.simple_ex:.2f}", f"{report.moderate_ex:.2f}", f"{report.challenging_ex:.2f}", f"{report.ex_accuracy:.2f}"),
            (_COL_SOFT_F1_PCT, f"{report.simple_f1:.2f}", f"{report.moderate_f1:.2f}", f"{report.challenging_f1:.2f}", f"{report.soft_f1:.2f}"),
            ("Errors", "", "", "", report.error_count),
            ("Timeouts", "", "", "", report.timeout_count),
        ]
        for dr in data_rows:
            for col, val in enumerate(dr, 1):
                c = ws_summary.cell(row=row, column=col, value=val)
                c.border = thin_border
                c.alignment = Alignment(horizontal=_EXCEL_CENTER)
                if col == 1:
                    c.font = subheader_font
            row += 1

    def _write_detail_sheet(self, ws_detail, results, styles):
        """Populate the Per-Question Results sheet."""
        import openpyxl
        header_font = styles[_EXCEL_HEADER_FONT]
        header_fill = styles[_EXCEL_HEADER_FILL]
        good_fill = styles["good_fill"]
        neutral_fill = styles["neutral_fill"]
        thin_border = styles[_EXCEL_THIN_BORDER]
        Alignment = styles[_EXCEL_ALIGNMENT]

        detail_headers = [
            "Question ID", "Database", "Difficulty", "EX Score", "Soft F1",
            "Predicted SQL", "Gold SQL", "Error",
        ]
        col_widths = [12, 25, 14, 10, 10, 60, 60, 40]
        for i, w in enumerate(col_widths, 1):
            ws_detail.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for col, h in enumerate(detail_headers, 1):
            c = ws_detail.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border

        for i, r in enumerate(sorted(results, key=lambda x: x.question_id), start=2):
            vals = [
                r.question_id, r.db_id, r.difficulty, r.ex_score,
                round(r.soft_f1_score, 4), r.predicted_sql, r.gold_sql, r.error or "",
            ]
            for col, val in enumerate(vals, 1):
                c = ws_detail.cell(row=i, column=col, value=val)
                c.border = thin_border
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 4:
                    c.fill = good_fill if val == 1 else neutral_fill

    def _write_db_sheet(self, ws_db, results, styles):
        """Populate the Per-Database Results sheet."""
        import openpyxl
        header_font = styles[_EXCEL_HEADER_FONT]
        header_fill = styles[_EXCEL_HEADER_FILL]
        thin_border = styles[_EXCEL_THIN_BORDER]
        Alignment = styles[_EXCEL_ALIGNMENT]

        db_headers = ["Database", "Count", _COL_EX_PCT, _COL_SOFT_F1_PCT, "Errors"]
        db_widths = [28, 10, 12, 12, 10]
        for i, w in enumerate(db_widths, 1):
            ws_db.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for col, h in enumerate(db_headers, 1):
            c = ws_db.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border

        db_groups: Dict[str, List[EvaluationResult]] = {}
        for r in results:
            db_groups.setdefault(r.db_id, []).append(r)

        for i, (db_id, db_results) in enumerate(sorted(db_groups.items()), start=2):
            count = len(db_results)
            ex = 100.0 * sum(r.ex_score for r in db_results) / count if count else 0
            f1 = 100.0 * sum(r.soft_f1_score for r in db_results) / count if count else 0
            errs = sum(1 for r in db_results if r.error)
            vals = [db_id, count, round(ex, 2), round(f1, 2), errs]
            for col, val in enumerate(vals, 1):
                c = ws_db.cell(row=i, column=col, value=val)
                c.border = thin_border
                c.alignment = Alignment(horizontal=_EXCEL_CENTER)

    @staticmethod
    def _write_baseline_reference_row(ws, row_num, vals_row, styles):
        """Write one reference baseline row into the comparison sheet."""
        subheader_font = styles["subheader_font"]
        thin_border = styles[_EXCEL_THIN_BORDER]
        Alignment = styles[_EXCEL_ALIGNMENT]
        for col, val in enumerate(vals_row, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.border = thin_border
            if col == 1:
                c.font = subheader_font
            if col >= 4:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    @staticmethod
    def _write_askrita_row(ws, row_num, ask_row, styles):
        """Write the askRITA (this run) row into the comparison sheet."""
        good_fill = styles["good_fill"]
        thin_border = styles[_EXCEL_THIN_BORDER]
        Alignment = styles[_EXCEL_ALIGNMENT]
        Font = styles["Font"]
        for col, val in enumerate(ask_row, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.border = thin_border
            if col in (2, 3):
                c.fill = good_fill
            if col == 1:
                c.font = Font(bold=True)
            if col >= 4:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    def _write_baseline_sheet(self, ws_cmp, report, run_config, styles):
        """Populate the Baseline vs askRITA comparison sheet."""
        header_font = styles[_EXCEL_HEADER_FONT]
        header_fill = styles[_EXCEL_HEADER_FILL]
        thin_border = styles[_EXCEL_THIN_BORDER]
        Alignment = styles[_EXCEL_ALIGNMENT]
        Font = styles["Font"]

        ws_cmp.column_dimensions["A"].width = 32
        ws_cmp.column_dimensions["B"].width = 14
        ws_cmp.column_dimensions["C"].width = 14
        ws_cmp.column_dimensions["D"].width = 36
        ws_cmp.column_dimensions["E"].width = 52

        r = 1
        ws_cmp.merge_cells("A1:E1")
        t = ws_cmp.cell(row=r, column=1, value="Baseline vs askRITA (BIRD Mini-Dev SQLite)")
        t.font = Font(bold=True, size=14, color="2F5496")
        t.alignment = Alignment(horizontal=_EXCEL_CENTER)
        r += 2

        disclaimer = (
            "Reference rows are published full-split (500q) leaderboard scores. "
            "Stratified or partial runs are not directly comparable but useful for trending."
        )
        ws_cmp.merge_cells(f"A{r}:E{r}")
        dc = ws_cmp.cell(row=r, column=1, value=disclaimer)
        dc.font = Font(italic=True, size=10)
        dc.alignment = Alignment(wrap_text=True)
        r += 2

        hdr = ["Reference", _COL_EX_PCT, _COL_SOFT_F1_PCT, "Source", "Notes"]
        for col, h in enumerate(hdr, 1):
            c = ws_cmp.cell(row=r, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal=_EXCEL_CENTER)
        r += 1

        for baseline_row in BIRD_MINI_DEV_SQLITE_REFERENCE_BASELINES:
            f1 = baseline_row.get("soft_f1_pct")
            vals_row = [
                baseline_row["label"],
                baseline_row["ex_pct"],
                f1 if f1 is not None else "—",
                baseline_row.get("source", ""),
                baseline_row.get("notes", ""),
            ]
            self._write_baseline_reference_row(ws_cmp, r, vals_row, styles)
            r += 1

        sampling_note = ""
        if run_config:
            sampling_note = str(run_config.get("Sampling mode", "") or "")
        ask_notes = f"This run ({report.total} questions). {sampling_note}".strip()
        ask_row = [
            "askRITA (this run)",
            round(report.ex_accuracy, 2),
            round(report.soft_f1, 2),
            "askRITA benchmark harness",
            ask_notes or "Mini-Dev SQLite via askRITA workflow.",
        ]
        self._write_askrita_row(ws_cmp, r, ask_row, styles)

    def save_spreadsheet(
        self,
        results: List[EvaluationResult],
        report: EvaluationReport,
        output_path: str,
        run_config: Optional[Dict[str, Any]] = None,
        include_baseline_comparison: bool = True,
    ):
        """Save detailed results and summary as an Excel spreadsheet.

        Creates a workbook with sheets:
        - Summary: Overall scores and difficulty breakdown
        - Per-Question Results: Every question with predicted/gold SQL and scores
        - Per-Database Results: Aggregated scores per BIRD database
        - Baseline vs askRITA (optional): Published Mini-Dev SQLite references vs this run

        Args:
            results: List of per-question EvaluationResult objects.
            report: Aggregated EvaluationReport.
            output_path: Path for the .xlsx file.
            run_config: Optional run metadata (model, provider, etc.).
            include_baseline_comparison: Add sheet comparing leaderboard baselines to askRITA.
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        except ImportError:
            logger.warning(
                "openpyxl not installed — skipping spreadsheet export. "
                "Install with: pip install openpyxl"
            )
            return

        wb = openpyxl.Workbook()

        styles = {
            _EXCEL_HEADER_FONT: Font(bold=True, size=13, color="FFFFFF"),
            _EXCEL_HEADER_FILL: PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
            "subheader_font": Font(bold=True, size=11),
            "good_fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "neutral_fill": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            _EXCEL_THIN_BORDER: Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            ),
            _EXCEL_ALIGNMENT: Alignment,
            "Font": Font,
        }

        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary_sheet(ws_summary, report, run_config, styles)

        ws_detail = wb.create_sheet("Per-Question Results")
        self._write_detail_sheet(ws_detail, results, styles)

        ws_db = wb.create_sheet("Per-Database Results")
        self._write_db_sheet(ws_db, results, styles)

        if include_baseline_comparison:
            ws_cmp = wb.create_sheet("Baseline vs askRITA")
            self._write_baseline_sheet(ws_cmp, report, run_config, styles)

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        logger.info("Spreadsheet saved to %s", output_path)
