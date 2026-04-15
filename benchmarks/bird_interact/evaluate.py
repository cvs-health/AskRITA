# Copyright 2026 CVS Health and/or one of its affiliates
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
# This file uses the following unmodified third-party packages,
# each retaining its original copyright and license:
#   openpyxl (MIT)

"""Evaluation module for BIRD Mini-Interact benchmark.

Implements the BIRD-Interact evaluation protocol:
  - Execute preprocess_sql (setup statements)
  - Run predicted SQL and compare against test cases
  - Execute clean_up_sql (teardown statements)
  - Reward scoring: 1.0 (first attempt), 0.5 (after debug retry), 0 (fail)

When ground-truth is not available, evaluation is skipped gracefully with
clear instructions on how to obtain it.
"""

import json
import logging
import os
import sqlite3
import time
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

from .runner import ConversationResult
from .setup_data import MiniInteractDataManager, MiniInteractTask

logger = logging.getLogger(__name__)

GT_REQUEST_INSTRUCTIONS = (
    "Ground-truth (GT) data is required for evaluation.\n"
    "To obtain it, email bird.bench25@gmail.com with subject:\n"
    "  [mini-interact GT&Test Cases]\n\n"
    "Once received, use --gt-path <path_to_gt.jsonl> to merge GT data\n"
    "and re-run evaluation with --evaluate-only."
)


def _execute_sql_safe(
    sql: str, db_path: str, timeout: float = 30.0
) -> Tuple[Optional[List[tuple]], Optional[str], float]:
    """Execute SQL against a SQLite database with timeout protection."""
    start = time.time()
    try:
        conn = sqlite3.connect(db_path, timeout=timeout)
        conn.execute("PRAGMA journal_mode=WAL;")
        cursor = conn.cursor()
        cursor.execute(sql)
        results = cursor.fetchall()
        conn.close()
        return results, None, time.time() - start
    except Exception as e:
        return None, str(e), time.time() - start


def _execute_statements(
    statements: List[str], db_path: str, timeout: float = 30.0
) -> Optional[str]:
    """Execute a list of SQL statements (preprocess / cleanup).

    Returns None on success, error string on failure.
    """
    try:
        conn = sqlite3.connect(db_path, timeout=timeout)
        conn.execute("PRAGMA journal_mode=WAL;")
        cursor = conn.cursor()
        for stmt in statements:
            stmt = stmt.strip()
            if stmt:
                cursor.execute(stmt)
        conn.commit()
        conn.close()
        return None
    except Exception as e:
        return str(e)


@dataclass
class TestCaseResult:
    """Result of evaluating a single test case."""

    test_case_id: str
    passed: bool
    expected: Any = None
    actual: Any = None
    error: Optional[str] = None


@dataclass
class TaskEvaluationResult:
    """Evaluation result for a single Mini-Interact task."""

    instance_id: str
    selected_database: str
    predicted_sql: str
    reward_score: float = 0.0
    test_cases_passed: int = 0
    test_cases_total: int = 0
    debug_used: bool = False
    error: Optional[str] = None
    test_case_results: List[TestCaseResult] = field(default_factory=list)

    @property
    def pass_rate(self) -> float:
        if self.test_cases_total == 0:
            return 0.0
        return self.test_cases_passed / self.test_cases_total


@dataclass
class InteractEvaluationReport:
    """Aggregated evaluation report for Mini-Interact benchmark."""

    total_tasks: int = 0
    evaluated_tasks: int = 0
    skipped_no_gt: int = 0

    overall_reward: float = 0.0
    avg_test_pass_rate: float = 0.0

    per_database: Dict[str, Dict[str, float]] = field(default_factory=dict)
    error_count: int = 0

    has_ground_truth: bool = False

    def to_dict(self) -> Dict[str, Any]:
        return {
            "total_tasks": self.total_tasks,
            "evaluated_tasks": self.evaluated_tasks,
            "skipped_no_gt": self.skipped_no_gt,
            "overall_reward": round(self.overall_reward, 4),
            "avg_test_pass_rate": round(self.avg_test_pass_rate, 4),
            "per_database": self.per_database,
            "error_count": self.error_count,
            "has_ground_truth": self.has_ground_truth,
        }

    def print_report(self) -> None:
        print("\n" + "=" * 80)
        print("BIRD Mini-Interact Evaluation Results")
        print("=" * 80)

        if not self.has_ground_truth:
            print("\n  [!] No ground-truth data available — evaluation skipped.")
            print(f"\n  {GT_REQUEST_INSTRUCTIONS}")
            print("=" * 80 + "\n")
            return

        print(f"  Total tasks:       {self.total_tasks}")
        print(f"  Evaluated:         {self.evaluated_tasks}")
        print(f"  Skipped (no GT):   {self.skipped_no_gt}")
        print(f"  Errors:            {self.error_count}")
        print("-" * 80)
        print(f"  Overall Reward:    {self.overall_reward:.4f}")
        print(f"  Avg Test Pass:     {self.avg_test_pass_rate:.4f}")
        print("-" * 80)

        if self.per_database:
            print(f"  {'Database':<30} {'Reward':>10} {'Tasks':>8}")
            print("  " + "-" * 50)
            for db_name in sorted(self.per_database.keys()):
                info = self.per_database[db_name]
                print(
                    f"  {db_name:<30} "
                    f"{info.get('reward', 0):.4f}     "
                    f"{info.get('count', 0):>5}"
                )

        print("=" * 80 + "\n")


@dataclass
class MiniInteractEvaluator:
    """Evaluates Mini-Interact benchmark results using test cases.

    When GT is unavailable, reports are generated without scores and
    the user is informed how to obtain GT data.

    Args:
        dataset_manager: MiniInteractDataManager with dataset access.
        timeout: SQL execution timeout per statement.
    """

    dataset_manager: MiniInteractDataManager
    timeout: float = 30.0

    def evaluate(
        self,
        conversation_results: List[ConversationResult],
        tasks: List[MiniInteractTask],
    ) -> Tuple[InteractEvaluationReport, List[TaskEvaluationResult]]:
        """Evaluate conversation results against ground-truth test cases.

        Args:
            conversation_results: Results from the runner.
            tasks: List of MiniInteractTask with GT fields.

        Returns:
            Tuple of (report, per-task evaluation results).
        """
        task_map = {t.instance_id: t for t in tasks}
        eval_results: List[TaskEvaluationResult] = []

        has_any_gt = any(t.has_ground_truth for t in tasks)

        report = InteractEvaluationReport(
            total_tasks=len(conversation_results),
            has_ground_truth=has_any_gt,
        )

        if not has_any_gt:
            logger.warning("No ground-truth data available — skipping evaluation.")
            report.skipped_no_gt = len(conversation_results)
            return report, eval_results

        for cr in conversation_results:
            task = task_map.get(cr.instance_id)
            if not task or not task.has_ground_truth:
                report.skipped_no_gt += 1
                continue

            result = self._evaluate_single(cr, task)
            eval_results.append(result)

        report.evaluated_tasks = len(eval_results)
        report.error_count = sum(1 for r in eval_results if r.error)

        if eval_results:
            report.overall_reward = sum(r.reward_score for r in eval_results) / len(
                eval_results
            )
            report.avg_test_pass_rate = sum(r.pass_rate for r in eval_results) / len(
                eval_results
            )

        # Per-database breakdown
        db_groups: Dict[str, List[TaskEvaluationResult]] = {}
        for r in eval_results:
            db_groups.setdefault(r.selected_database, []).append(r)

        for db_name, db_results in db_groups.items():
            count = len(db_results)
            avg_reward = sum(r.reward_score for r in db_results) / count
            report.per_database[db_name] = {
                "count": count,
                "reward": round(avg_reward, 4),
            }

        return report, eval_results

    def _evaluate_single(
        self, cr: ConversationResult, task: MiniInteractTask
    ) -> TaskEvaluationResult:
        """Evaluate a single conversation result against GT.

        Strategy:
        - If test_cases are available: run them (the BIRD protocol).
        - If only sol_sql is available: fall back to execution accuracy (EX),
          comparing predicted vs gold result sets, same as BIRD Mini-Dev.
        """
        db_path = self.dataset_manager.get_db_path(task.selected_database)
        test_cases = task.test_cases or []

        result = TaskEvaluationResult(
            instance_id=cr.instance_id,
            selected_database=cr.selected_database,
            predicted_sql=cr.predicted_sql,
            test_cases_total=len(test_cases),
            debug_used=cr.debug_used,
        )

        # Run preprocess SQL
        preprocess = task.preprocess_sql or []
        if preprocess:
            err = _execute_statements(preprocess, db_path, self.timeout)
            if err:
                result.error = f"Preprocess failed: {err}"
                self._cleanup(task, db_path)
                return result

        try:
            pred_results, pred_err, _ = _execute_sql_safe(
                cr.predicted_sql, db_path, self.timeout
            )
            if pred_err:
                result.error = f"Predicted SQL error: {pred_err}"
                self._cleanup(task, db_path)
                return result

            if test_cases:
                self._score_via_test_cases(
                    result, test_cases, pred_results, db_path, cr
                )
            elif task.sol_sql:
                self._score_via_execution_accuracy(
                    result, pred_results, task.sol_sql, db_path, cr
                )

        except Exception as e:
            result.error = str(e)
        finally:
            self._cleanup(task, db_path)

        return result

    def _score_via_test_cases(
        self,
        result: TaskEvaluationResult,
        test_cases: List[Dict[str, Any]],
        pred_results: Optional[List[tuple]],
        db_path: str,
        cr: ConversationResult,
    ) -> None:
        """Score using BIRD test-case protocol."""
        tc_results = self._run_test_cases(test_cases, pred_results, db_path)
        result.test_case_results = tc_results
        result.test_cases_passed = sum(1 for tc in tc_results if tc.passed)
        all_passed = (
            result.test_cases_total > 0
            and result.test_cases_passed == result.test_cases_total
        )
        if all_passed:
            result.reward_score = 0.5 if cr.debug_used else 1.0
        else:
            result.reward_score = 0.0

    def _score_via_execution_accuracy(
        self,
        result: TaskEvaluationResult,
        pred_results: Optional[List[tuple]],
        gold_sql: str,
        db_path: str,
        cr: ConversationResult,
    ) -> None:
        """Score by comparing predicted vs gold SQL execution results (EX)."""
        gold_results, gold_err, _ = _execute_sql_safe(gold_sql, db_path, self.timeout)
        if gold_err:
            result.error = f"Gold SQL error: {gold_err}"
            return
        match = self._compare_results(pred_results, gold_results)
        result.test_cases_total = 1
        result.test_cases_passed = 1 if match else 0
        if match:
            result.reward_score = 0.5 if cr.debug_used else 1.0
        else:
            result.reward_score = 0.0

    def _run_test_cases(
        self,
        test_cases: List[Dict[str, Any]],
        pred_results: Optional[List[tuple]],
        db_path: str,
    ) -> List[TestCaseResult]:
        """Evaluate predicted results against test cases."""
        return [
            self._run_single_test_case(idx, tc, pred_results, db_path)
            for idx, tc in enumerate(test_cases)
        ]

    def _run_single_test_case(
        self,
        idx: int,
        tc: Dict[str, Any],
        pred_results: Optional[List[tuple]],
        db_path: str,
    ) -> TestCaseResult:
        """Evaluate a single test case."""
        tc_id = tc.get("test_id", f"tc_{idx}")
        tc_type = tc.get("type", "result_match")
        tc_sql = tc.get("sql", "")
        tc_expected = tc.get("expected")
        actual_str = str(pred_results)[:200] if pred_results else None

        if tc_type == "result_match" and tc_sql:
            return self._eval_result_match(tc_id, tc_sql, pred_results, db_path)
        if tc_type == "value_check" and tc_expected is not None:
            passed = self._check_value_in_results(pred_results, tc_expected)
            return TestCaseResult(
                test_case_id=tc_id,
                passed=passed,
                expected=str(tc_expected),
                actual=actual_str,
            )
        return TestCaseResult(
            test_case_id=tc_id,
            passed=False,
            error=f"Unknown test case type: {tc_type}",
        )

    def _eval_result_match(
        self,
        tc_id: str,
        tc_sql: str,
        pred_results: Optional[List[tuple]],
        db_path: str,
    ) -> TestCaseResult:
        """Evaluate a result_match test case by executing SQL and comparing."""
        expected_res, err, _ = _execute_sql_safe(tc_sql, db_path, self.timeout)
        if err:
            return TestCaseResult(
                test_case_id=tc_id,
                passed=False,
                error=f"Test case SQL error: {err}",
            )
        passed = self._compare_results(pred_results, expected_res)
        return TestCaseResult(
            test_case_id=tc_id,
            passed=passed,
            expected=str(expected_res)[:200] if expected_res else None,
            actual=str(pred_results)[:200] if pred_results else None,
        )

    def _compare_results(
        self,
        actual: Optional[List[tuple]],
        expected: Optional[List[tuple]],
    ) -> bool:
        """Compare two result sets (order-insensitive)."""
        if actual is None or expected is None:
            return actual is None and expected is None
        return set(actual) == set(expected)

    def _check_value_in_results(
        self,
        results: Optional[List[tuple]],
        expected_value: Any,
    ) -> bool:
        """Check if an expected value appears anywhere in the results."""
        if results is None:
            return False
        expected_str = str(expected_value)
        for row in results:
            for cell in row:
                if str(cell) == expected_str:
                    return True
        return False

    def _cleanup(self, task: MiniInteractTask, db_path: str) -> None:
        """Execute clean_up_sql statements to restore DB state."""
        cleanup = task.clean_up_sql or []
        if cleanup:
            err = _execute_statements(cleanup, db_path, self.timeout)
            if err:
                logger.warning("Cleanup failed for task %s: %s", task.instance_id, err)

    def save_report(
        self,
        report: InteractEvaluationReport,
        output_path: str,
    ) -> None:
        """Save evaluation report to a JSON file."""
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        with open(output_path, "w") as f:
            json.dump(report.to_dict(), f, indent=2)
        logger.info("Evaluation report saved to %s", output_path)

    def save_spreadsheet(
        self,
        eval_results: List[TaskEvaluationResult],
        conversation_results: List[ConversationResult],
        report: InteractEvaluationReport,
        output_path: str,
        run_config: Optional[Dict[str, Any]] = None,
    ) -> None:
        """Save results as an Excel spreadsheet.

        Sheets:
          - Summary: Overall scores and run config
          - Per-Task Results: Detailed per-task evaluation
          - Conversations: Full conversation history
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            logger.warning(
                "openpyxl not installed — skipping spreadsheet export. "
                "Install with: pip install openpyxl"
            )
            return

        wb = openpyxl.Workbook()
        styles = {
            "header_font": Font(bold=True, size=13, color="FFFFFF"),
            "header_fill": PatternFill(
                start_color="2F5496", end_color="2F5496", fill_type="solid"
            ),
            "good_fill": PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            ),
            "neutral_fill": PatternFill(
                start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
            ),
            "thin_border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
            "Alignment": Alignment,
            "Font": Font,
        }

        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary_sheet(ws_summary, report, run_config, styles)

        ws_tasks = wb.create_sheet("Per-Task Results")
        self._write_tasks_sheet(ws_tasks, eval_results, styles)

        ws_conv = wb.create_sheet("Conversations")
        self._write_conversations_sheet(ws_conv, conversation_results, styles)

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        logger.info("Spreadsheet saved to %s", output_path)

    def _write_summary_sheet(self, ws, report, run_config, styles):
        """Populate the Summary sheet."""
        font_cls = styles["Font"]
        align_cls = styles["Alignment"]

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25

        row = 1
        ws.merge_cells("A1:B1")
        cell = ws.cell(
            row=row,
            column=1,
            value="BIRD Mini-Interact — askRITA Evaluation Report",
        )
        cell.font = font_cls(bold=True, size=16, color="2F5496")
        cell.alignment = align_cls(horizontal="center")
        row += 2

        if run_config:
            for key, val in run_config.items():
                ws.cell(row=row, column=1, value=key).font = font_cls(
                    bold=True, size=11
                )
                ws.cell(row=row, column=2, value=str(val))
                row += 1
            row += 1

        metrics = [
            ("Total Tasks", report.total_tasks),
            ("Evaluated", report.evaluated_tasks),
            ("Skipped (no GT)", report.skipped_no_gt),
            ("Errors", report.error_count),
            ("Overall Reward", round(report.overall_reward, 4)),
            ("Avg Test Pass Rate", round(report.avg_test_pass_rate, 4)),
        ]
        for label, value in metrics:
            ws.cell(row=row, column=1, value=label).font = font_cls(bold=True, size=11)
            ws.cell(row=row, column=2, value=value)
            row += 1

        if not report.has_ground_truth:
            row += 1
            ws.merge_cells(f"A{row}:B{row}")
            gt_cell = ws.cell(row=row, column=1, value=GT_REQUEST_INSTRUCTIONS)
            gt_cell.alignment = align_cls(wrap_text=True)

    def _write_tasks_sheet(self, ws, eval_results, styles):
        """Populate the Per-Task Results sheet."""
        import openpyxl

        header_font = styles["header_font"]
        header_fill = styles["header_fill"]
        thin_border = styles["thin_border"]
        good_fill = styles["good_fill"]
        neutral_fill = styles["neutral_fill"]
        align_cls = styles["Alignment"]

        task_headers = [
            "Instance ID",
            "Database",
            "Predicted SQL",
            "Reward",
            "Tests Passed",
            "Tests Total",
            "Pass Rate",
            "Debug Used",
            "Error",
        ]
        col_widths = [20, 25, 60, 10, 12, 12, 12, 10, 40]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for col, h in enumerate(task_headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border

        for i, r in enumerate(eval_results, start=2):
            vals = [
                r.instance_id,
                r.selected_database,
                r.predicted_sql,
                r.reward_score,
                r.test_cases_passed,
                r.test_cases_total,
                round(r.pass_rate, 4),
                r.debug_used,
                r.error or "",
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=i, column=col, value=val)
                c.border = thin_border
                c.alignment = align_cls(wrap_text=True, vertical="top")
                if col == 4:
                    c.fill = good_fill if val >= 0.5 else neutral_fill

    def _write_conversations_sheet(self, ws, conversation_results, styles):
        """Populate the Conversations sheet."""
        import openpyxl

        header_font = styles["header_font"]
        header_fill = styles["header_fill"]
        thin_border = styles["thin_border"]
        align_cls = styles["Alignment"]

        conv_headers = [
            "Instance ID",
            "Database",
            "Query",
            "Turns",
            "SQL Generated",
            "Latency (s)",
            "Conversation",
        ]
        conv_widths = [20, 20, 50, 8, 60, 12, 80]
        for i, w in enumerate(conv_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for col, h in enumerate(conv_headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border

        for row_idx, cr in enumerate(conversation_results, start=2):
            conv_text = "\n".join(
                f"[{t.role.upper()} T{t.turn_number}] {t.content[:200]}"
                for t in cr.conversation_turns
            )
            vals = [
                cr.instance_id,
                cr.selected_database,
                cr.amb_user_query,
                cr.num_turns,
                cr.predicted_sql,
                round(cr.latency_seconds, 2),
                conv_text,
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.border = thin_border
                c.alignment = align_cls(wrap_text=True, vertical="top")
